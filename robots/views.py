from django.views import View
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.db.models import Count

from datetime import timedelta
import os
import itertools
import tempfile
import json
from openpyxl import Workbook

from .forms import RobotForm
from .models import Robot


class PostJson(View):
    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        form = RobotForm(data)
        if form.is_valid():
            form.save()
            return JsonResponse({'message': 'Data received successfully'})
        errors = form.errors.as_join()
        return JsonResponse({'errors': errors}, status=400)


class RobotsExcelView(View):
    def get(self, request, *args, **kwargs):
        # Generate report within report_delta
        report_delta = {'weeks': 1}

        # TODO: Feat: Synchronize timezone
        report_time_end = timezone.now()
        report_time_start = report_time_end - timedelta(**report_delta)

        robots = Robot.objects.filter(
            created__gte=report_time_start
        ).values('model', 'version').annotate(
            num_counted=Count('id', distinct=True)
        ).order_by('model', 'version')

        grouped_robots = {}
        for key, group in itertools.groupby(robots, lambda k: k["model"]):
            grouped_robots[key] = list(group)

        ROBOT_COL_TITLES = [
            {
                "key": "model",
                "name": "Модель",
                "col_num": 1
            },
            {
                "key": "version",
                "name": "Версия",
                "col_num": 2
            },
            {
                "key": "num_counted",
                "name": "Количество за неделю",
                "col_num": 3
            },
        ]

        report_settings = {
            "filename": "wk-rep-robots",
            "report_time": report_time_end,
            "titles": ROBOT_COL_TITLES,
            "grouped_data": grouped_robots,
        }

        temp_fp = make_excel_report(report_settings)

        try:
            return FileResponse(temp_fp)
        except Exception as error:
            raise error
        finally:
            os.remove(temp_fp)
            os.rmdir(os.path.dirname(tempfile))


def make_excel_report(**settings):
    report_wb = Workbook()
    report_wb.remove(report_wb.active)

    report_filename = settings["filename"]
    report_time = settings["report_time"]
    report_date = report_time.strftime('%d.%m.%Y'),

    temp_dir = tempfile.mkdtemp(dir=os.path.dirname(__file__))
    temp_fp = os.path.join(
        temp_dir, f"{report_filename}-{report_date}.xlsx"
    )

    col_titles = settings["titles"]

    for model, group in settings["grouped_data"].items():
        report_ws = report_wb.create_sheet(model)
        set_ws_column_titles(report_ws, col_titles)
        set_ws_data(report_ws, group, col_titles)

    report_wb.save(temp_fp)
    return temp_fp


def set_ws_column_titles(ws, col_titles):
    for title in col_titles:
        ws.cell(
            row=1,
            column=title["col_num"],
            value=title["name"],
        )


def set_ws_data(ws, data, col_titles):
    current_row = 2
    for item in data:
        for title in col_titles:
            ws.cell(
                row=current_row,
                column=title["col_num"],
                value=item[title["key"]],
            )
        current_row += 1
